import datetime
import json
import os
from pipes import Template
import re
import openpyxl
import requests

def read(file, is_enhance):
    # 读取Excel文件
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active
    result = []
    for row in sheet.iter_rows():
        if len(row[0].value) != 6:
            continue
        if is_enhance and '增强' in str(row[1].value) or \
                not is_enhance and '增强' not in str(row[1].value):
            result.append({'code':row[0].value,
                           'name':row[1].value,
                           'date':str(row[9].value)})
    return result

def read_gem(file, is_enhance):
    # 读取Excel文件
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active
    result = []
    for row in sheet.iter_rows():
        if len(row[7].value) != 6 or '境外' in row[5].value:
            continue
        if is_enhance and '增强' in str(row[1].value) or \
                not is_enhance and '增强' not in str(row[1].value):
            result.append({'code':row[7].value,
                           'name':row[1].value,
                           'date':str(row[2].value).replace(' 00:00:00', '')})
    return result

def get_fund_list(is_enhance):
    sz50 = read('source/上证50.xlsx', is_enhance)
    hs300 = read('source/沪深300.xlsx', is_enhance)
    zz500 = read('source/中证500.xlsx', is_enhance)
    zz1000 = read('source/中证1000.xlsx', is_enhance)
    cyb = read_gem('source/创业板指.xlsx', is_enhance)
    return sz50 + hs300 + zz500 + zz1000 + cyb


def get_fund_manager(data):
    for item in data:
        item['manager_code'] = request_manager_code(item['code'])
    return data

def request_manager_code(fund_code):
    cache = read_manager_cache()
    manager_code = 0
    if fund_code not in cache:
        url_template = 'https://fundf10.eastmoney.com/jjjl_{fund_code}.html'.format(fund_code=fund_code)
        page = requests.get(url_template, headers={
                "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                            'Chrome/78.0.3904.108 Safari/537.36'})
        pattern = re.compile(r'基金经理：&nbsp;&nbsp;<a href="//fund.eastmoney.com/manager/(.+?).html')
        result = pattern.search(page.text)
        if result:
            manager_code = result.group(1)
            cache[fund_code] = manager_code
            write_manager_cache(cache)
    else:
        manager_code = cache[fund_code]
    return manager_code

def read_manager_cache():
    manager_cache = {}
    if os.path.exists('manager_cache.json'):
        with open('manager_cache.json', 'r') as f:
            manager_cache = json.load(f)
    return manager_cache

def write_manager_cache(cache):
    # 将字典保存到JSON文件中
    with open('manager_cache.json', 'w') as f:
        json.dump(cache, f)